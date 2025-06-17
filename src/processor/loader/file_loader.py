from pathlib import Path
from typing import Dict, Any
import pandas as pd
import xlrd
import openpyxl

from .base import BaseLoader


class FileLoader(BaseLoader):
    """
    ファイルローダー
    CSV、XLS、XLSXの全ファイル形式に対応
    """
    
    def get_supported_extensions(self) -> set:
        return {".csv", ".xls", ".xlsx"}
    
    def load_file_internal(self, file_path: Path) -> Dict[str, Any]:
        """
        ファイル形式を自動判定して適切な読み込み処理を実行
        """
        extension = file_path.suffix.lower()
        
        if extension == ".csv":
            return self._load_csv(file_path)
        elif extension == ".xls":
            return self._load_xls(file_path)
        elif extension == ".xlsx":
            return self._load_xlsx(file_path)
        else:
            raise ValueError(f"サポートされていないファイル形式: {extension}")
    
    def _load_csv(self, file_path: Path) -> Dict[str, Any]:
        """CSV形式の読み込み処理"""
        try:
            try:
                df = pd.read_csv(file_path, header=None, encoding="utf-8")
            except UnicodeDecodeError:
                df = pd.read_csv(file_path, header=None, encoding="cp932")
            self.logger.info(f"CSV読み込み完了: shape={df.shape}")
        except UnicodeDecodeError as e:
            self.logger.error(f"CSVのエンコーディングエラー: {e}")
            raise ValueError("CSVファイルのエンコーディングに問題があります。Shift_JIS（cp932）などを試してください。")
        
        sheet_info = self.create_sheet_info("CSV", df)
        return self.create_result_structure(file_path, ".csv", [sheet_info])
    
    def _load_xls(self, file_path: Path) -> Dict[str, Any]:
        """XLS形式の読み込み処理"""
        self.logger.info(f"=== .xls ファイル読み込み開始: {file_path} ===")
        
        sheets = []
        try:
            wb = xlrd.open_workbook(str(file_path), formatting_info=True)
            self.logger.info(f"xlrdで開いたワークブック: {wb.nsheets} シート")

            for sheet in wb.sheets():
                self.logger.info(f"=== シート '{sheet.name}' を処理中 ===")
                rows = []
                for row_idx in range(sheet.nrows):
                    row = sheet.row_values(row_idx)
                    rows.append(row)

                df = pd.DataFrame(rows)
                self.logger.info(f"xlrdで構築したDataFrame: shape={df.shape}")

                sheet_info = self.create_sheet_info(sheet.name, df)
                sheets.append(sheet_info)

        except Exception as e:
            self.logger.error(f".xls読み込みでエラー: {e}")
            self.logger.exception("詳細なエラー情報:")
            # エラー時は空のデータフレームを返す
            sheet_info = self.create_sheet_info("Sheet1", pd.DataFrame())
            sheets.append(sheet_info)
        
        return self.create_result_structure(file_path, ".xls", sheets)
    
    def _load_xlsx(self, file_path: Path) -> Dict[str, Any]:
        """XLSX形式の読み込み処理"""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheets = []
        
        for ws in wb.worksheets:
            try:
                # ヘッダーは後で LLM が判定するため、ここでは header=None
                df = pd.read_excel(file_path, sheet_name=ws.title, header=None)
                self.logger.info(f"xlsx シート '{ws.title}' 読み込み完了: shape={df.shape}")
            except Exception as e:
                self.logger.error(f"xlsx シート '{ws.title}' の読み込みでエラー: {e}")
                # 読み込みに失敗した場合は空 DataFrame
                df = pd.DataFrame()

            sheet_info = self.create_sheet_info(ws.title, df)
            sheets.append(sheet_info)
        
        return self.create_result_structure(file_path, ".xlsx", sheets) 