import os
import re
import zipfile
import tempfile
import xml.etree.ElementTree as ET
import pandas as pd
from typing import Optional, Tuple
from openpyxl.workbook.workbook import Workbook
from .utils import (
    detect_platform_characters,
    detect_notes_outside_table,
    get_excel_column_letter
)

def check_no_images_or_objects(
    df: Optional[pd.DataFrame] = None,
    workbook: Optional[Workbook] = None,
    filepath: Optional[str] = None
) -> Tuple[bool, str]:
    if not filepath or not zipfile.is_zipfile(filepath):
        return False, "有効な Excel ファイルではありません"

    with zipfile.ZipFile(filepath, 'r') as zip_ref:
        with tempfile.TemporaryDirectory() as tmpdir:
            zip_ref.extractall(tmpdir)

            drawing_files = []
            for root, _, files in os.walk(tmpdir):
                for file in files:
                    if file.startswith("drawing") and file.endswith(".xml"):
                        drawing_files.append(os.path.join(root, file))

            if not drawing_files:
                return True, "図形・テキストボックスは見つかりませんでした"

            found_shapes = []
            for path in drawing_files:
                try:
                    tree = ET.parse(path)
                    root = tree.getroot()
                    xml_str = ET.tostring(root, encoding='unicode').lower()
                    if "textbox" in xml_str or "shape" in xml_str or "<a:txbody>" in xml_str:
                        found_shapes.append(os.path.basename(path))
                except Exception as e:
                    return False, f"XML解析中にエラー: {e}"

            if found_shapes:
                return False, f"図形・テキストボックスが検出されました: {found_shapes}"
            return True, "図形・テキストボックスは見つかりませんでした"

def check_no_format_based_semantics(
    df: Optional[pd.DataFrame] = None,
    workbook: Optional[Workbook] = None,
    filepath: Optional[str] = None
) -> Tuple[bool, str]:
    if workbook is None:
        return False, "エラー: 有効な workbook が渡されていません"

    styled_cells = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb != '00000000':
                    styled_cells.append(cell.coordinate)

    if styled_cells:
        return False, f"書式による強調セルがあります（例: {styled_cells[:3]})"
    return True, "書式ベースの意味づけは検出されませんでした"

def check_no_merged_cells(
    df: Optional[pd.DataFrame] = None,
    workbook: Optional[Workbook] = None,
    filepath: Optional[str] = None
) -> Tuple[bool, str]:
    if workbook is None:
        return False, "エラー: 有効な workbook が渡されていません"

    merged = []
    for sheet in workbook.worksheets:
        merged += [str(rng) for rng in sheet.merged_cells.ranges]

    if merged:
        return False, f"結合セルが検出されました: {merged}"
    return True, "結合セルはありません"

def check_valid_file_format(
    df: Optional[pd.DataFrame] = None,
    workbook: Optional[Workbook] = None,
    filepath: Optional[str] = None
) -> Tuple[bool, str]:
    if not filepath:
        return False, "ファイルパスが指定されていません"

    ext = os.path.splitext(filepath)[1].lower()
    if ext not in ['.csv', '.xlsx']:
        return False, f"サポート外のファイル形式です: {ext}"
    return True, "ファイル形式はCSVまたはExcelです"

def check_single_data_per_cell(
    df: Optional[pd.DataFrame] = None,
    workbook: Optional[Workbook] = None,
    filepath: Optional[str] = None
) -> Tuple[bool, str]:
    if workbook is None:
        return False, "エラー: 有効な workbook が渡されていません"

    problem_cells = []
    pattern = re.compile(r"[\n,;/]")
    sheet = workbook.worksheets[0]

    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        for col_idx, cell in enumerate(row, start=1):
            if isinstance(cell, str) and pattern.search(cell):
                col_letter = get_excel_column_letter(col_idx)
                cell_pos = f"{col_letter}{row_idx}"
                problem_cells.append((cell_pos, cell))

    if problem_cells:
        examples = [f"{pos}: '{val}'" for pos, val in problem_cells[:3]]
        return False, f"複数データセルが検出されました（例: {examples}）"

    return True, "各セルに1データのみです"

def check_no_whitespace_formatting(
    df: Optional[pd.DataFrame] = None,
    workbook: Optional[Workbook] = None,
    filepath: Optional[str] = None
) -> Tuple[bool, str]:
    if workbook is None:
        return False, "エラー: 有効な workbook が渡されていません"

    problem = []
    sheet = workbook.worksheets[0]

    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        for col_idx, cell in enumerate(row, start=1):
            if isinstance(cell, str):
                trimmed = cell.strip()
                if re.fullmatch(r"[¥￥]?\d+(,?\d+)*(円|万円)?", trimmed):
                    continue
                if cell != trimmed or '\n' in cell or '\t' in cell or re.search(r"\w　\w", cell):
                    col_letter = get_excel_column_letter(col_idx)
                    cell_pos = f"{col_letter}{row_idx}"
                    problem.append(f"{cell_pos}: {repr(cell)}")

    if problem:
        return False, f"余分な空白/改行/体裁スペースが検出されました（例: {problem[:3]}）"
    return True, "スペースや改行による整形はありません"

def check_one_table_per_sheet(
    df: Optional[pd.DataFrame] = None,
    workbook: Optional[Workbook] = None,
    filepath: Optional[str] = None
) -> Tuple[bool, str]:
    if workbook is None:
        return False, "エラー: 有効な workbook が渡されていません"

    issues = []

    for sheet in workbook.worksheets:
        data_flags = [
            any(cell not in (None, "") for cell in row)
            for row in sheet.iter_rows(values_only=True)
        ]

        in_table = False
        block_count = 0
        for has_data in data_flags:
            if has_data and not in_table:
                block_count += 1
                in_table = True
            elif not has_data:
                in_table = False

        if block_count > 1:
            issues.append(sheet.title)

    if issues:
        return False, f"複数の表が含まれている可能性のあるシート: {issues}"
    return True, "各シートに1つの表のみです"

def check_no_hidden_rows_or_columns(
    df: Optional[pd.DataFrame] = None,
    workbook: Optional[Workbook] = None,
    filepath: Optional[str] = None
) -> Tuple[bool, str]:
    if workbook is None:
        return False, "エラー: 有効な workbook が渡されていません"

    issues = []
    for sheet in workbook.worksheets:
        hidden_rows = [dim.index for dim in sheet.row_dimensions.values() if dim.hidden]
        hidden_cols = [dim.index for dim in sheet.column_dimensions.values() if dim.hidden]
        if hidden_rows or hidden_cols:
            issues.append({
                'sheet': sheet.title,
                'hidden_rows': hidden_rows,
                'hidden_cols': hidden_cols
            })

    if issues:
        return False, f"非表示行/列があります: {issues}"
    return True, "行や列の非表示はありません"

def check_no_notes_outside_table(
    df: Optional[pd.DataFrame] = None,
    workbook: Optional[Workbook] = None,
    filepath: Optional[str] = None
) -> Tuple[bool, str]:
    if workbook is None:
        return False, "エラー: 有効な workbook が渡されていません"

    texts = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
            for cell in row:
                if isinstance(cell, str) and cell.strip():
                    texts.append(cell.strip())

    if not texts:
        return True, "表外のメモや備考は検出されませんでした"

    result = detect_notes_outside_table(texts)
    if "注釈あり" in result:
        return False, "表外に備考/注釈が含まれている可能性があります"
    return True, "表外の注釈や備考は含まれていません"

def check_no_platform_dependent_characters(
    df: Optional[pd.DataFrame] = None,
    workbook: Optional[Workbook] = None,
    filepath: Optional[str] = None
) -> Tuple[bool, str]:
    if workbook is None:
        return False, "エラー: 有効な workbook が渡されていません"

    issues = []
    sheet = workbook.worksheets[0]

    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        for col_idx, cell in enumerate(row, start=1):
            if isinstance(cell, str) and detect_platform_characters(cell):
                col_letter = get_excel_column_letter(col_idx)
                cell_pos = f"{col_letter}{row_idx}"
                issues.append(f"{cell_pos}: '{cell}'")

    if issues:
        return False, f"機種依存文字が含まれています（例: {issues[:3]}）"
    return True, "機種依存文字は含まれていません"
