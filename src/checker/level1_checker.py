from pathlib import Path
import re
from typing import Tuple, cast, Any
from openpyxl.workbook.workbook import Workbook
import pandas as pd
import xlrd
import zipfile
from loguru import logger

from src.processor.context import TableContext

logger.add("logs/checker1.log", rotation="10 MB", retention="30 days", level="DEBUG")


def _col_to_num(col_str: str) -> int:
    """Excelの列文字（A, B, AAなど）を数値に変換する"""
    num = 0
    for char in col_str:
        num = num * 26 + (ord(char.upper()) - ord("A")) + 1
    return num


def get_sort_key(cell_str: str) -> Tuple[int, int]:
    """セル番地文字列からソートキー（行番号, 列番号）を返す"""
    # 代表的な形式: "A1", "A1:...", "A1（...", "列A 行12: ..."
    m = re.search(r"([A-Z]+)(\d+)", cell_str)
    if m:
        col, row = m.groups()
        return (int(row), _col_to_num(col))
    m2 = re.search(r"列\s*([A-Z]+)\s*行\s*(\d+)", cell_str)
    if m2:
        col, row = m2.groups()
        return (int(row), _col_to_num(col))
    return (99999, 99999)


def get_excel_column_letter(n: int) -> str:
    result = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result


def get_xls_workbook_info(file_path: Path) -> dict:
    """xlsファイルの基本情報を取得"""
    try:
        workbook = xlrd.open_workbook(str(file_path))

        sheet_info = []
        for sheet_name in workbook.sheet_names():
            sheet = workbook.sheet_by_name(sheet_name)
            sheet_info.append(
                {"name": sheet_name, "nrows": sheet.nrows, "ncols": sheet.ncols}
            )

        return {
            "file_path": str(file_path),
            "nsheets": workbook.nsheets,
            "sheet_names": workbook.sheet_names(),
            "sheet_info": sheet_info,
        }
    except Exception as e:
        logger.error(f"xlsファイルの詳細情報取得でエラー: {e}")
        return {}


def has_any_drawing(path: Path) -> bool:
    """
    Excel ファイルに図形やオブジェクトが含まれているかをチェック
    .xls ファイルは構造上チェックが困難なため常に False を返す
    """
    ext = path.suffix.lower()
    if ext == ".xls":
        # .xls ファイルは構造上図形チェックが困難なため、
        # 図形があるものとして扱う（必要に応じて後で対応）
        return True
    elif ext != ".xlsx":
        return False

    try:
        with zipfile.ZipFile(path, "r") as z:
            for name in z.namelist():
                if name.startswith("xl/drawings/") and name.endswith(".xml"):
                    xml = z.read(name)
                    if b"<xdr:twoCellAnchor" in xml or b"<xdr:oneCellAnchor" in xml:
                        return True
    except Exception:
        return False
    return False


def detect_platform_characters(text: str) -> bool:
    pattern = re.compile(r"[①-⑳⓪-⓿Ⅰ-Ⅻ㊤㊥㊦㊧㊨㈱㈲㈹℡〒〓※]")
    return bool(pattern.search(text))


def is_clean_numeric(val: Any) -> bool:
    if isinstance(val, (int, float)):
        return True
    if isinstance(val, str):
        s = val.strip()
        if re.search(r"[^\d.\-]", s):
            return False
        try:
            float(s)
            return True
        except ValueError:
            return False
    return False


FREE_TEXT_PATTERN = re.compile(
    r"""
    ^\s*(?:  
        (?:その他|そのほか)\s*[:：\-\–\/]           |
        (?:その他|そのほか)\s*[\(（].+?[\)）]       |
        コメント\s*[:：]                            |
        自由記述\s*[:：]                            |
        詳細\s*[:：]                                |
        備考\s*[:：]                                |
        補足\s*[:：]                                |
        感想\s*[:：]                                |
        意見\s*[:：]                                |
        メモ\s*[:：]                                |
        特記事項\s*[:：]                            |
        注釈\s*[:：]                                |
        自己PR\s*[:：]                              |
        フリーテキスト\s*[:：]                      |
        フリー回答\s*[:：]
    )
""",
    re.VERBOSE,
)


MISSING_VALUE_EXPRESSIONS = [
    "不明",
    "不詳",
    "…",
    "無記入",
    "無回答",
    "該当なし",
    "なし",
    "無し",
    "n/a",
    "na",
    "nan",
    "未定",
    "未記入",
    "未入力",
    "未回答",
    "記載なし",
    "対象外",
    "空欄",
    "空白",
    "不在",
    "特になし",
    "---",
    "--",
    "-",
    "ー",
    "―",
    "？",
    "?",
    "わからない",
    "わかりません",
    "なし（特記なし）",
    "無し（詳細不明）",
    "無効",
    "省略",
    "null",
    "none",
]


def is_likely_long_format(df: pd.DataFrame) -> bool:
    """
    ID・変数名・値 を含み、列数が多い DataFrame を縦型とみなす。
    """
    if len(df.columns) < 10:
        return False
    return {"ID", "変数名", "値"}.issubset(set(df.columns))


def check_xls_merged_cells(
    file_path: Path, sheet_name: str, row_start: int, row_end: int
) -> list:
    """xlsファイルの結合セルをチェック（指定範囲内のみ）"""
    try:
        wb = xlrd.open_workbook(str(file_path), formatting_info=True)
        sheet = wb.sheet_by_name(sheet_name)
        merged_cells = []

        # xlrdのmerged_cellsは (r0, r1, c0, c1) 形式。r1/c1 は「1つ先」のインデックス。
        for r0, r1, c0, c1 in getattr(sheet, "merged_cells", []):
            last_row = r1 - 1
            # テーブル本体の行範囲内かどうか
            if r0 >= row_start and last_row <= row_end:
                top_left = f"{get_excel_column_letter(c0 + 1)}{r0 + 1}"
                bottom_right = f"{get_excel_column_letter(c1)}{r1}"
                merged_cells.append(f"{top_left}:{bottom_right}")

        return merged_cells

    except Exception as e:
        logger.error(f"check_xls_merged_cells エラー ({sheet_name}): {e}")
        return []


def check_xls_cell_formats(
    file_path: Path, sheet_name: str, data_start: int, data_end: int
) -> list:
    """xlsファイルのセル書式をチェック（修正版）"""
    try:
        logger.debug(f"check_xls_cell_formats: 開始 - {file_path}, sheet: {sheet_name}")
        workbook = xlrd.open_workbook(str(file_path), formatting_info=True)
        sheet = workbook.sheet_by_name(sheet_name)
        flagged = []

        for row_idx in range(data_start, min(data_end + 1, sheet.nrows)):
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                xf_index = cell.xf_index

                if xf_index >= len(workbook.xf_list):
                    continue  # 異常なインデックスはスキップ

                xf = workbook.xf_list[xf_index]
                font_index = xf.font_index

                if font_index >= len(workbook.font_list):
                    continue

                font = workbook.font_list[font_index]
                coord = f"{get_excel_column_letter(col_idx + 1)}{row_idx + 1}"

                # 太字
                if font.bold:
                    flagged.append(f"{coord}（太字）")
                # イタリック
                if font.italic:
                    flagged.append(f"{coord}（イタリック）")
                # 下線
                if font.underline_type != 0:
                    flagged.append(f"{coord}（下線）")
                # 文字色
                if font.colour_index not in (0, 1, 7, 8):  # 自動・黒・白以外
                    flagged.append(f"{coord}（文字色）")
                # 背景色
                bg_index = xf.background.pattern_colour_index
                if bg_index not in (64, 0):  # 標準色以外
                    flagged.append(f"{coord}（背景色）")

        return flagged

    except Exception as e:
        logger.exception(f"書式チェックでエラー: {e}")
        return []


def detect_multiple_tables_dataframe(
    df: pd.DataFrame, sheet_name: str = "", data_start_offset: int = 0
) -> tuple:
    """
    DataFrameベースで複数テーブルを検出する

    Args:
        df: 対象のDataFrame
        sheet_name: シート名（ログ用）

    Returns:
        tuple: (has_multiple_tables: bool, details: str)
    """
    try:
        if df.empty or len(df) < 3:
            return False, "データが少ないため複数テーブルの検出をスキップ"

        # 完全に空の行を検索
        empty_rows = []
        for idx, row in df.iterrows():
            if row.isna().all() or (row.astype(str).str.strip() == "").all():
                empty_rows.append(idx)

        # 連続する空行を検索（テーブル区切りの可能性）
        if len(empty_rows) > 0:
            consecutive_groups = []
            current_group = [empty_rows[0]]

            for i in range(1, len(empty_rows)):
                if empty_rows[i] == empty_rows[i - 1] + 1:
                    current_group.append(empty_rows[i])
                else:
                    if len(current_group) >= 2:  # 2行以上の連続空行
                        consecutive_groups.append(current_group)
                    current_group = [empty_rows[i]]

            if len(current_group) >= 2:
                consecutive_groups.append(current_group)

            if consecutive_groups:
                return (
                    True,
                    f"複数の連続空行グループが見つかりました: {len(consecutive_groups)}箇所",
                )

        # ヘッダー様の行の検出
        header_like_rows = []
        for idx, row in df.iterrows():
            non_na_values = row.dropna().astype(str).str.strip()
            if len(non_na_values) > 0:
                # 数値以外が多い行をヘッダー候補とする
                numeric_count = sum(
                    1
                    for val in non_na_values
                    if val.replace(".", "").replace("-", "").isdigit()
                )
                if numeric_count / len(non_na_values) < 0.5:
                    # 元ファイルの実行番号（0-based index + offset + 1）に変換
                    actual_row_number = idx + data_start_offset + 1
                    header_like_rows.append(actual_row_number)

        # 複数のヘッダー様行が離れて存在する場合
        if len(header_like_rows) >= 2:
            gaps = [
                header_like_rows[i + 1] - header_like_rows[i]
                for i in range(len(header_like_rows) - 1)
            ]
            if any(gap > 3 for gap in gaps):  # 3行以上離れたヘッダーがある
                return (
                    True,
                    f"離れた位置に複数のヘッダー様行が検出されました: {header_like_rows}",
                )

        return False, "単一テーブルと判定"

    except Exception as e:
        logger.error(f"複数テーブル検出でエラー: {e}")
        return False, f"検出処理でエラーが発生: {str(e)}"


def check_valid_file_format(
    ctx: TableContext, workbook: Workbook, filepath: str
) -> Tuple[bool, str]:
    ext = Path(filepath).suffix.lower()
    if ext not in {".csv", ".xlsx", ".xls"}:
        return False, f"サポート外のファイル形式です: {ext}"
    elif ext == ".xls":
        return (
            True,
            "旧Excel（.xls）形式のため、一部の自動チェック（書式・図形など）が制限されます。必要に応じて目視での確認を行ってください",
        )
    return True, "ファイル形式はCSVまたはExcel（.xlsx）です"


def check_no_images_or_objects(
    ctx: TableContext, workbook: Workbook, filepath: str
) -> Tuple[bool, str]:
    path = Path(filepath)
    ext = path.suffix.lower()
    if ext == ".csv":
        return True, "csvファイルのためオブジェクトチェック不要"
    elif ext == ".xls":
        # .xlsファイルでは図形・オブジェクトの詳細検出は困難だが、
        # 一般的に統計表では図形は使用されないため、警告として扱う
        return (
            False,
            "xlsファイルでは図形や画像の自動判定ができません。必要に応じて目視でご確認ください",
        )
    elif ext == ".xlsx":
        if has_any_drawing(path):
            return False, "図形・テキストボックスが検出されました"
        return True, "図形・テキストボックスは見つかりませんでした"
    else:
        return True, "サポート外形式のためオブジェクトチェック不要"


def check_one_table_per_sheet(
    ctx: TableContext, workbook: Workbook, filepath: str
) -> Tuple[bool, str]:
    # .xlsファイルの場合はDataFrameベースでチェック
    if workbook is None:
        # 元のDataFrameを使用して複数テーブル検出（行番号はデータ開始行のオフセットを考慮）
        data_start = ctx.row_indices.get("data_start", 0)
        is_multiple, details = detect_multiple_tables_dataframe(
            ctx.data, ctx.sheet_name, data_start_offset=data_start
        )

        if is_multiple:
            return False, f"複数テーブルの疑いがあります: {details}"
        return True, "1つのテーブルのみです"

    ws = workbook[ctx.sheet_name]
    column_rows = ctx.row_indices.get("column_rows")
    data_end = ctx.row_indices.get("data_end")

    if column_rows is None or data_end is None:
        return False, "シート範囲情報が不足しているためチェックできません"

    start = (
        min(column_rows) if isinstance(column_rows, list) else cast(int, column_rows)
    )
    end = cast(int, data_end)

    flags = [
        any(cell not in (None, "") for cell in row)
        for row in ws.iter_rows(min_row=start + 1, max_row=end + 1, values_only=True)
    ]

    in_block = False
    blocks = 0
    for has_data in flags:
        if has_data and not in_block:
            blocks += 1
            in_block = True
        elif not has_data:
            in_block = False

    if blocks > 1:
        return False, f"複数テーブルの疑いがあります（検出ブロック数: {blocks}）"
    return True, "1つのテーブルのみです"


def check_xls_hidden_rows_columns(file_path: Path) -> tuple:
    """xlsファイルの非表示行・列をチェック（修正版）"""
    try:
        logger.debug(f"check_xls_hidden_rows_columns: 開始 - {file_path}")
        workbook = xlrd.open_workbook(str(file_path), formatting_info=True)
        hidden_rows = []
        hidden_cols = []

        for sheet_name in workbook.sheet_names():
            logger.debug(f"シート処理中: {sheet_name}")
            sheet = workbook.sheet_by_name(sheet_name)

            # 行の高さが0 → 非表示
            for row_idx in range(sheet.nrows):
                rowinfo = sheet.rowinfo_map.get(row_idx)
                if rowinfo:
                    logger.debug(f"  row {row_idx}: height={rowinfo.height}")
                if rowinfo and rowinfo.height == 0:
                    logger.info(f"  非表示行検出: {sheet_name} 行{row_idx}")
                    hidden_rows.append((sheet_name, row_idx))

            # 列の幅が0 → 非表示（colinfo_map は sheet 単位）
            for col_idx, colinfo in sheet.colinfo_map.items():
                logger.debug(f"  col {col_idx}: width={colinfo.width}")
                if colinfo.width == 0:
                    logger.info(f"  非表示列検出: {sheet_name} 列{col_idx}")
                    hidden_cols.append((sheet_name, col_idx))

        return hidden_rows, hidden_cols

    except Exception as e:
        logger.exception(f"非表示行・列チェックでエラー: {e}")
        return [], []


def check_no_hidden_rows_or_columns(
    ctx: TableContext, workbook: Workbook, filepath: str
) -> Tuple[bool, str]:
    # ブックが無い場合（.xls または .csv など）
    if workbook is None:
        ext = Path(filepath).suffix.lower()
        if ext == ".xls":
            hidden_rows, hidden_cols = check_xls_hidden_rows_columns(Path(filepath))

            row_str = (
                ", ".join(f"{sheet}シートの{r + 1}行" for sheet, r in hidden_rows)
                if hidden_rows
                else "該当なし"
            )
            col_str = (
                ", ".join(
                    f"{sheet}シートの{get_excel_column_letter(c + 1)}列"
                    for sheet, c in hidden_cols
                )
                if hidden_cols
                else "該当なし"
            )

            if hidden_rows or hidden_cols:
                return False, f"非表示行／列があります（行: {row_str}, 列: {col_str}）"
            return True, "非表示行／列はありません"
        elif ext == ".csv":
            return True, "csvファイルのため非表示行・列の概念はありません"
        else:
            return True, "サポート外形式のため非表示行・列チェックをスキップします"

    # .xlsx の場合
    ws = workbook[ctx.sheet_name]
    hidden_rows = [d.index for d in ws.row_dimensions.values() if d.hidden]
    hidden_cols = [d.index for d in ws.column_dimensions.values() if d.hidden]

    row_str = ", ".join(f"{r}行" for r in hidden_rows) if hidden_rows else "該当なし"
    col_str = (
        ", ".join(f"{get_excel_column_letter(c)}列" for c in hidden_cols)
        if hidden_cols
        else "該当なし"
    )

    if hidden_rows or hidden_cols:
        return False, f"非表示行／列があります（行: {row_str}, 列: {col_str}）"
    return True, "非表示行／列はありません"


def check_no_notes_outside_table(
    ctx: TableContext, workbook: Workbook, filepath: str
) -> Tuple[bool, str]:
    problem_notes: list[str] = []

    # 上部注釈
    if not ctx.upper_annotations.empty:
        for row_idx, row in ctx.upper_annotations.iterrows():
            actual_row = row_idx + 1
            content = [str(v) for v in row.dropna().values]
            if content:
                problem_notes.append(f"**{actual_row}行目:** {', '.join(content)}")

    # 下部注釈
    if not ctx.lower_annotations.empty:
        for row_idx, row in ctx.lower_annotations.iterrows():
            actual_row = row_idx + 1
            content = [str(v) for v in row.dropna().values]
            if content:
                problem_notes.append(f"**{actual_row}行目:** {', '.join(content)}")

    if problem_notes:
        details = "\n- ".join(problem_notes)
        return False, f"テーブルの範囲外で以下の内容が検出されました:\n- {details}"

    return True, "表外の注釈や備考はありません"


def check_no_merged_cells(
    ctx: TableContext, workbook: Workbook, filepath: str
) -> Tuple[bool, str]:
    # .xls は workbook=None になるためこちら
    if workbook is None:
        ext = Path(filepath).suffix.lower()
        if ext == ".xls":
            # テーブル本体の開始・終了行を ctx から取得
            column_rows = ctx.row_indices.get("column_rows")
            data_end = ctx.row_indices.get("data_end")
            if column_rows is None or data_end is None:
                return False, "結合セルチェックに必要な情報が不足しています"

            start = (
                min(column_rows)
                if isinstance(column_rows, list)
                else cast(int, column_rows)
            )
            end = cast(int, data_end)

            merged = check_xls_merged_cells(Path(filepath), ctx.sheet_name, start, end)
            if merged:
                return False, f"結合セルが検出されました: {merged}"
            else:
                return True, "結合セルはありません"
        elif ext == ".csv":
            return True, "csvファイルのため結合セルは存在しません"
        else:
            return True, "サポート外形式のため結合セルチェックをスキップします"

    ws = workbook[ctx.sheet_name]
    column_rows = ctx.row_indices.get("column_rows")
    data_end = ctx.row_indices.get("data_end")

    if column_rows is None or data_end is None:
        return False, "結合セルチェックに必要な情報が不足しています"

    start = (
        min(column_rows) + 1
        if isinstance(column_rows, list)
        else cast(int, column_rows) + 1
    )
    end = cast(int, data_end) + 1

    relevant_merges = [
        str(rng)
        for rng in ws.merged_cells.ranges
        if rng.min_row >= start and rng.max_row <= end
    ]

    if relevant_merges:
        relevant_merges.sort(key=get_sort_key)
        return False, f"結合セルが検出されました: {relevant_merges}"
    return True, "結合セルはありません"


def check_no_format_based_semantics(
    ctx: TableContext, workbook: Workbook, filepath: str
) -> Tuple[bool, str]:
    # 先頭でファイル拡張子を確認し、.xls は早期に警告付き合格とする
    path = Path(filepath)
    ext = path.suffix.lower()

    if ext == ".xls":
        return (
            True,
            "旧Excel（.xls）形式のため、書式（文字色や太字など）の自動判定が不正確な場合があります。必要に応じて目視でご確認ください。",
        )
    # .xlsファイルの場合はxlrdを使用してチェック
    if workbook is None:
        ext = Path(filepath).suffix.lower()
        if ext == ".xls":
            data_start = ctx.row_indices.get("data_start", 0)
            data_end = ctx.row_indices.get("data_end", len(ctx.data) - 1)

            flagged = check_xls_cell_formats(
                Path(filepath), ctx.sheet_name, data_start, data_end
            )

            if flagged:
                flagged.sort(key=get_sort_key)
                details = "\n- ".join(flagged)
                return False, f"以下のセルで視覚的装飾による意味付けが検出されました:\n- {details}"
            return True, "書式ベースの意味づけは検出されませんでした"
        elif ext == ".csv":
            return True, "csvファイルのため書式装飾チェックは対象外です"
        else:
            return True, "サポート外形式のため書式装飾チェックをスキップします"

    ws = workbook[ctx.sheet_name]
    column_rows = ctx.row_indices.get("column_rows")
    data_end = ctx.row_indices.get("data_end")

    if column_rows is None or data_end is None:
        return False, "書式チェックに必要な情報が不足しています"

    start = (
        min(column_rows) + 1
        if isinstance(column_rows, list)
        else cast(int, column_rows) + 1
    )
    end = cast(int, data_end) + 1

    flagged = []
    for row in ws.iter_rows(min_row=start, max_row=end):
        for cell in row:
            coord = cell.coordinate
            fill = cell.fill
            if fill and fill.fgColor:
                fg = fill.fgColor
                if hasattr(fg, "rgb") and isinstance(fg.rgb, str):
                    rgb = fg.rgb.upper()
                    if rgb not in ("00000000", "FFFFFFFF", "FF000000"):
                        flagged.append(f"{coord}（塗りつぶし）")

            font = cell.font
            if font:
                if (
                    font.color
                    and hasattr(font.color, "rgb")
                    and isinstance(font.color.rgb, str)
                ):
                    rgb = font.color.rgb.upper()
                    if rgb not in ("00000000", "FF000000"):
                        flagged.append(f"{coord}（文字色）")

                if font.bold:
                    flagged.append(f"{coord}（太字）")
                if font.italic:
                    flagged.append(f"{coord}（イタリック）")
                if font.underline:
                    flagged.append(f"{coord}（下線）")
                if font.sz and (font.sz < 9 or font.sz > 13):
                    flagged.append(f"{coord}（フォントサイズ {font.sz}）")

    if flagged:
        flagged.sort(key=get_sort_key)
        details = "\n- ".join(flagged)
        return False, f"以下のセルで視覚的装飾による意味付けが検出されました:\n- {details}"
    return True, "書式ベースの意味づけは検出されませんでした"


def check_no_whitespace_formatting(
    ctx: TableContext, workbook: Workbook, filepath: str
) -> Tuple[bool, str]:
    # .xlsファイルの場合はワークブックがNoneになるため、DataFrameベースでチェック
    if workbook is None:
        # DataFrameを使用した簡易版の空白チェック
        sample_cells = []
        for row_idx, row in ctx.data.iterrows():
            for col_idx, val in enumerate(row):
                if isinstance(val, str) and "　" in val:
                    col_letter = get_excel_column_letter(col_idx + 1)
                    cell_ref = f"{col_letter}{row_idx + 1}"
                    sample_cells.append(f"{cell_ref}: '{val.strip()}'")
                    if len(sample_cells) >= 10:
                        break
            if len(sample_cells) >= 10:
                break

        if not sample_cells:
            return True, "体裁調整目的の空白は見つかりませんでした"

        # 全角スペースが見つかった時点で不合格とする
        sample_cells.sort(key=get_sort_key)
        details = "\n- ".join(sample_cells)
        return False, f"以下のセルで体裁調整目的の空白が使用されている可能性があります:\n- {details}"

    ws = workbook[ctx.sheet_name]
    column_rows = ctx.row_indices.get("column_rows")
    data_end = ctx.row_indices.get("data_end")

    if column_rows is None or data_end is None:
        return False, "空白チェックに必要な情報が不足しています"

    start = (
        min(column_rows) + 1
        if isinstance(column_rows, list)
        else cast(int, column_rows) + 1
    )
    end = cast(int, data_end) + 1

    sample_cells = []
    for r_idx, row in enumerate(
        ws.iter_rows(min_row=start, max_row=end, values_only=True), start=start
    ):
        for c_idx, val in enumerate(row, start=1):
            if isinstance(val, str) and "　" in val:
                col_letter = get_excel_column_letter(c_idx)
                cell_ref = f"{col_letter}{r_idx}"
                sample_cells.append(f"{cell_ref}: '{val.strip()}'")
                if len(sample_cells) >= 10:
                    break
        if len(sample_cells) >= 10:
            break

    if not sample_cells:
        return True, "体裁調整目的の空白は見つかりませんでした"

    # 全角スペースが見つかった時点で不合格とする
    sample_cells.sort(key=get_sort_key)
    details = "\n- ".join(sample_cells)
    return False, f"以下のセルで体裁調整目的の空白が使用されている可能性があります:\n- {details}"


def check_single_data_per_cell(
    ctx: TableContext, workbook: Workbook, filepath: str
) -> Tuple[bool, str]:
    pattern = re.compile(r"[\n,;/]")
    problems = []

    data_start = ctx.row_indices.get("data_start")
    if data_start is None:
        return False, "データ開始位置が不明です"

    start: int = cast(int, data_start)
    for row_idx_raw, row in ctx.data.iterrows():
        row_idx: int = cast(int, row_idx_raw)
        for col_idx, val in enumerate(row):
            if isinstance(val, str) and pattern.search(val):
                excel_row = row_idx + 1 + start
                excel_col_letter = get_excel_column_letter(col_idx + 1)
                coord = f"{excel_col_letter}{excel_row}"
                problems.append(f"{coord}: {repr(val)}")

    if problems:
        problems.sort(key=get_sort_key)
        details = "\n- ".join(problems)
        return False, f"以下のセルで複数のデータが検出されました:\n- {details}"
    return True, "各セルに1データのみです"


def check_no_platform_dependent_characters(
    ctx: TableContext, workbook: Workbook, filepath: str
) -> Tuple[bool, str]:
    # .xlsファイルの場合はワークブックがNoneになるため、DataFrameベースでチェック
    if workbook is None:
        # DataFrameを使用した簡易版の機種依存文字チェック
        issues = []
        for row_idx, row in ctx.data.iterrows():
            for col_idx, val in enumerate(row):
                if isinstance(val, str) and detect_platform_characters(val):
                    coord = f"{get_excel_column_letter(col_idx + 1)}{row_idx + 1}"
                    issues.append(f"{coord}: '{val}'")

        if issues:
            issues.sort(key=get_sort_key)
            details = "\n- ".join(issues)
            return False, f"以下のセルで機種依存文字が含まれています:\n- {details}"
        return True, "機種依存文字は含まれていません"

    ws = workbook[ctx.sheet_name]
    column_rows = ctx.row_indices.get("column_rows")
    data_end = ctx.row_indices.get("data_end")

    if column_rows is None or data_end is None:
        return False, "機種依存文字チェックに必要な情報が不足しています"

    if isinstance(column_rows, list):
        start = min(column_rows) + 1
    else:
        start = cast(int, column_rows) + 1
    end = cast(int, data_end) + 1

    issues = []
    for r, row in enumerate(
        ws.iter_rows(min_row=start, max_row=end, values_only=True), start=start
    ):
        for c, val in enumerate(row, start=1):
            if isinstance(val, str) and detect_platform_characters(val):
                coord = f"{get_excel_column_letter(c)}{r}"
                issues.append(f"{coord}: '{val}'")

    if issues:
        issues.sort(key=get_sort_key)
        details = "\n- ".join(issues)
        return False, f"以下のセルで機種依存文字が含まれています:\n- {details}"
    return True, "機種依存文字は含まれていません"


def check_numeric_columns_only(
    ctx: TableContext, workbook: Workbook, filepath: str
) -> Tuple[bool, str]:
    df = ctx.data
    problem_cells: dict = {}

    for col_idx, col_name in enumerate(df.columns):
        series_raw = df.iloc[:, col_idx]
        # 重複カラムで DataFrame になる場合はフラット化
        if isinstance(series_raw, pd.DataFrame):
            series = series_raw.stack(dropna=True).reset_index(drop=True)
        else:
            series = series_raw.dropna()
        if series.empty:
            continue

        total = len(series)
        ok_count = series.apply(is_clean_numeric).sum()
        if ok_count / total < 0.8:
            # 数値列とはみなさない
            continue

        if ok_count / total < 0.99:
            for row_idx, val in zip(series.index, series):
                if not is_clean_numeric(val):
                    coord = f"{get_excel_column_letter(col_idx + 1)}{row_idx + 1}"
                    problem_cells.setdefault(col_name, []).append(f"{coord}: '{val}'")

    if problem_cells:
        for cells in problem_cells.values():
            cells.sort(key=get_sort_key)
        msgs = [
            f"{name}:\n  - " + "\n  - ".join(cells) for name, cells in problem_cells.items()
        ]
        return False, "数値列に数値以外が含まれています:\n" + "\n".join(msgs)

    return True, "数値列に不正なデータは含まれていません"


def check_separate_other_detail_columns(
    ctx: TableContext, workbook: Workbook, filepath: str
) -> Tuple[bool, str]:
    df = ctx.data
    flagged = []

    for col_idx, col in enumerate(df.columns, start=1):
        col_series = df[col]
        if isinstance(col_series, pd.DataFrame):
            series = col_series.stack(dropna=True).reset_index(drop=True).astype(str)
        else:
            if not pd.api.types.is_string_dtype(col_series):
                continue
            series = col_series.dropna().astype(str)
        if series.empty:
            continue

        if series.str.contains(FREE_TEXT_PATTERN).any():
            flagged.append(f"{col}（列: {get_excel_column_letter(col_idx)}）")

    if flagged:
        return False, f"選択肢列に自由記述が混在している可能性があります: {flagged}"
    return True, "選択肢列と自由記述は適切に分離されています"


def check_no_missing_column_headers(
    ctx: TableContext, workbook: Workbook, filepath: str
) -> Tuple[bool, str]:
    df = ctx.data
    suspect = [c for c in df.columns if "Unnamed" in str(c) or str(c).strip() == ""]

    def is_unclear(name: str) -> bool:
        s = name.strip()
        if s == "":
            return True
        # 1文字だけ、または記号/数字のみは不明瞭とみなす
        if len(s) <= 1:
            return True
        if re.fullmatch(r"\d+", s):
            return True
        if re.fullmatch(r"[\W_]+", s):
            return True
        # 例: A, B1, X2 など短い略号を不明瞭とみなす
        if re.fullmatch(r"[A-Za-z](\d)?", s):
            return True
        return False

    for col in df.columns:
        if col in suspect:
            continue
        if is_unclear(str(col)):
            suspect.append(col)

    if suspect:
        return False, f"省略・不明な列名が検出されました: {suspect}"
    return True, "全ての列に意味のあるヘッダーが付いています"


def check_handling_of_missing_values(
    ctx: TableContext, workbook: Workbook, filepath: str
) -> Tuple[bool, str]:
    df = ctx.data
    # 問題を値ごとにグループ化するための辞書
    problems_by_value: dict[str, list[str]] = {}

    data_start_offset = ctx.row_indices.get("data_start", 0)
    normalized_missing = {str(x).strip().lower() for x in MISSING_VALUE_EXPRESSIONS}

    for col_idx, col_name in enumerate(df.columns):
        col_series = df.iloc[:, col_idx]

        for row_idx, val in col_series.items():
            if isinstance(val, str):
                cleaned_val = val.strip()
                if cleaned_val.lower() in normalized_missing:
                    excel_row = row_idx + data_start_offset + 1
                    excel_col = get_excel_column_letter(col_idx + 1)
                    cell_coord = f"{excel_col}{excel_row}"

                    if cleaned_val not in problems_by_value:
                        problems_by_value[cleaned_val] = []
                    problems_by_value[cleaned_val].append(cell_coord)

    if problems_by_value:
        # グループ化された結果から、整形されたメッセージを生成
        message_parts = ["以下のセルで欠損値を示す表現が見つかりました:"]
        for value, cells in sorted(problems_by_value.items()):
            cell_list_str = ", ".join(cells)
            message_parts.append(
                f"  - **値『{value}』** が見つかったセル ({len(cells)}件): {cell_list_str}"
            )

        details = "\n".join(message_parts)
        return False, details

    return True, "欠損表現は検出されませんでした"


def check_csv_single_line_per_data(
    ctx: TableContext, workbook: Workbook, filepath: str
) -> Tuple[bool, str]:
    ext = Path(filepath).suffix.lower()
    if ext != ".csv":
        return True, "CSVファイルではないためチェック対象外"

    df = ctx.data.copy()

    problem_cells = []
    for col_idx, col_name in enumerate(df.columns):
        col_series = df.iloc[:, col_idx]

        # 文字列型以外はスキップ
        if not pd.api.types.is_string_dtype(col_series) and not pd.api.types.is_object_dtype(col_series):
            continue

        if isinstance(col_series, pd.DataFrame):
            series = col_series.stack(dropna=True, future_stack=True).astype(str)
        else:
            series = col_series.dropna().astype(str)
        new_line_problems = series[series.str.contains(r"[\n\r]", na=False)]

        if not new_line_problems.empty:
            for row_idx_raw, val in new_line_problems.items():
                row_idx = row_idx_raw[0] if isinstance(row_idx_raw, tuple) else row_idx_raw
                coord = f"列{get_excel_column_letter(col_idx + 1)} 行{row_idx + 1}"
                display_val = str(val).replace("\n", "↵").replace("\r", "↵")
                problem_cells.append(f"{coord}: '{display_val[:20]}...'")

    if problem_cells:
        problem_cells.sort(key=get_sort_key)
        details = "\n- ".join(problem_cells)
        return False, f"以下のセルでデータ内部に改行が含まれています:\n- {details}"
    return True, "データ内部に改行は含まれていません"


def check_csv_fields_quoted(
    ctx: TableContext, workbook: Workbook, filepath: str
) -> Tuple[bool, str]:
    ext = Path(filepath).suffix.lower()
    if ext != ".csv":
        return True, "CSVファイルではないためチェック対象外"

    try:
        with open(filepath, "r", encoding="utf-8") as f:
            content = f.read()
    except Exception:
        try:
            with open(filepath, "r", encoding="shift_jis") as f:
                content = f.read()
        except Exception as e:
            return False, f"ファイル読み込みエラー: {e}"

    has_unquoted_comma = False

    for line in content.splitlines():
        fields = line.split(",")
        if any(
            "," in field
            and not (field.strip().startswith('"') and field.strip().endswith('"'))
            for field in fields
        ):
            has_unquoted_comma = True
            break

    if has_unquoted_comma:
        return (
            False,
            "カンマを含むフィールドがダブルクォーテーションで囲まれていない可能性があります。データが正しく列分割されないリスクがあります。",
        )

    return (
        True,
        "フィールドは適切に囲まれているか、またはカンマを含まないことが推測されます",
    )


CHECK_FUNCTIONS = {
    "check_valid_file_format": check_valid_file_format,
    "check_no_images_or_objects": check_no_images_or_objects,
    "check_one_table_per_sheet": check_one_table_per_sheet,
    "check_no_hidden_rows_or_columns": check_no_hidden_rows_or_columns,
    "check_no_notes_outside_table": check_no_notes_outside_table,
    "check_no_merged_cells": check_no_merged_cells,
    "check_no_format_based_semantics": check_no_format_based_semantics,
    "check_no_whitespace_formatting": check_no_whitespace_formatting,
    "check_single_data_per_cell": check_single_data_per_cell,
    "check_no_platform_dependent_characters": check_no_platform_dependent_characters,
    "check_no_missing_column_headers": check_no_missing_column_headers,
    "check_handling_of_missing_values": check_handling_of_missing_values,
    "check_csv_single_line_per_data": check_csv_single_line_per_data,
    "check_csv_fields_quoted": check_csv_fields_quoted,
}
